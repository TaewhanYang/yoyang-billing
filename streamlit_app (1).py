
import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import os

st.set_page_config(page_title="요양원 청구 마법사", layout="wide")

st.title("🏥 요양원 청구 마법사")
st.write("두 개의 엑셀 파일을 업로드하면 자동으로 요양원별 청구 정리가 됩니다.")

with st.expander("📎 예시파일 다운로드"):
    try:
        with open("examples/조제데이터_raw_예시.xlsx", "rb") as f:
            st.download_button("⬇️ 조제데이터_raw 예시.xlsx", data=f, file_name="조제데이터_raw_예시.xlsx")
        with open("examples/요양원기본테이블_예시.xlsx", "rb") as f:
            st.download_button("⬇️ 요양원기본테이블 예시.xlsx", data=f, file_name="요양원기본테이블_예시.xlsx")
    except FileNotFoundError:
        st.warning("예시 파일을 찾을 수 없습니다. 'examples' 폴더가 존재해야 합니다.")

uploaded_raw = st.file_uploader("1️⃣ 조제 데이터 (요양원 없는 원본)", type=["xlsx"])
uploaded_base = st.file_uploader("2️⃣ 요양원 기본 테이블", type=["xlsx"])

if uploaded_raw and uploaded_base:
    raw_df = pd.read_excel(uploaded_raw)
    base_df = pd.read_excel(uploaded_base)

    if "고객이름+주민등록번호" not in raw_df.columns and "고객이름" in raw_df.columns and "주민등록번호" in raw_df.columns:
        raw_df["고객이름+주민등록번호"] = raw_df["고객이름"].astype(str) + raw_df["주민등록번호"].astype(str)

    if "고객이름+주민등록번호" not in base_df.columns and "고객이름" in base_df.columns and "주민등록번호" in base_df.columns:
        base_df["고객이름+주민등록번호"] = base_df["고객이름"].astype(str) + base_df["주민등록번호"].astype(str)

    st.subheader("👀 업로드한 데이터 미리보기")
    st.write("🔹 조제데이터 원본 (자동 컬럼 생성 포함):")
    st.dataframe(raw_df.head(10), use_container_width=True)

    st.write("🔹 요양원기본테이블:")
    st.dataframe(base_df.head(10), use_container_width=True)

    merged_df = pd.merge(
        raw_df,
        base_df[["고객이름+주민등록번호", "요양원명"]],
        on="고객이름+주민등록번호",
        how="left",
        suffixes=("", "_기본")
    )
    merged_df["요양원명"] = merged_df["요양원명"].fillna(merged_df["요양원명_기본"])
    merged_df = merged_df.drop(columns=["요양원명_기본"])

    st.write("✅ 요양원 자동입력 결과:")
    st.dataframe(merged_df.head(10), use_container_width=True)

    # 전체 리스트 다운로드
    full_output = BytesIO()
    merged_df.to_excel(full_output, index=False)
    st.download_button(
        label="📥 전체 자동입력 리스트 다운로드",
        data=full_output.getvalue(),
        file_name="전체리스트_요양원자동입력.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    if st.button("🚀 요양원별 청구 엑셀 생성"):
        grouped = merged_df.groupby("요양원명")
        wb = Workbook()
        ws_index = wb.active
        ws_index.title = "목차"
        ws_index["A1"] = "요양원 청구 리스트"
        ws_index["A1"].font = Font(bold=True, size=14)
        index_row = 3

        for name, group in grouped:
            sheet_name = name[:31]
            ws = wb.create_sheet(title=sheet_name)
            for r in dataframe_to_rows(group, index=False, header=True):
                ws.append(r)
            if "요양급여액" in group.columns and "계" in group.columns:
                last_row = ws.max_row + 1
                ws[f"A{last_row}"] = "합계"
                ws[f"A{last_row}"].font = Font(bold=True)
                col1 = get_column_letter(group.columns.get_loc("요양급여액")+1)
                col2 = get_column_letter(group.columns.get_loc("계")+1)
                ws[f"{col1}{last_row}"] = f"=SUM({col1}2:{col1}{last_row-1})"
                ws[f"{col2}{last_row}"] = f"=SUM({col2}2:{col2}{last_row-1})"
            cell = ws_index.cell(row=index_row, column=1, value=name)
            cell.font = Font(color="0000FF", underline="single")
            cell.hyperlink = f"#{sheet_name}!A1"
            index_row += 1

        wb.active = wb["목차"]
        output = BytesIO()
        wb.save(output)
        st.success("✅ 요양원별 청구파일이 완성되었습니다. 아래에서 다운로드하세요.")
        st.download_button(
            label="📥 요양원별 청구파일 다운로드",
            data=output.getvalue(),
            file_name="요양원별_청구파일_자동생성.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
